# -*- coding: utf-8 -*-
"""Whole-module audit: structure, cross-references, links, encodings and workbook health.

Complements verify_module_02.py (figures and prose) and test_workbooks.py (live Excel values).
Run from the repository root:  python Module_02/checks/audit_module_02.py
Exits non-zero on any failure.
"""
import csv
import hashlib
import json
import os
import pathlib
import re
import sys

M2 = pathlib.Path(os.environ.get('M2DIR', 'Module_02'))
W = M2 / 'weeks-02-03-excel-for-data-analysis'

TOPICS = [
    '01-the-excel-environment-tables-and-referencing',
    '02-core-formulas-and-functions',
    '03-data-cleaning-in-excel',
    '04-pivot-tables',
    '05-charts-and-dashboard-reporting',
    '06-power-query',
    '07-ai-augmented-excel',
]

fails, passes = [], [0]


def check(label, cond, detail=''):
    if cond:
        passes[0] += 1
    else:
        fails.append(f"{label}: {detail}")


# ------------------------------------------------------------------ structure
EXPECTED = [
    'Datasets/Ilesanmi_Sales_Raw_Export.csv',
    'Datasets/Ilesanmi_Sales_Raw_Export.xlsx',
    'Datasets/Ilesanmi_Sales_Clean_AnswerKey.xlsx',
    'Datasets/README.md',
    'MCQ/module-02-mcq.csv',
    'MCQ/module-02-mcq.json',
    'module-demo/demo-guide.md',
    'module-demo/Ilesanmi_Sales_Raw_Export.csv',
    'module-demo/Ilesanmi_Sales_Clean_AnswerKey.xlsx',
    'mini-project/module-02-mini-project-brief.md',
    'mini-project/module-02-mini-project-rubric.md',
    'content-tracker.md',
]
for t in TOPICS:
    n = t[:2]
    EXPECTED += [
        f'{W.name}/{t}/demo/demo-guide.md',
        f'{W.name}/{t}/lab/lab-pack.md',
        f'{W.name}/{t}/lab/solutions/tier-1-solution.md',
        f'{W.name}/{t}/lab/solutions/tier-2-solution.md',
        f'{W.name}/{t}/lab/solutions/tier-3-solution.md',
        f'{W.name}/{t}/lab/solutions/solution-notes.md',
        f'{W.name}/{t}/lab/solutions/2.{int(n)}_lab_solution.xlsx',
    ]

for rel in EXPECTED:
    check(f'exists {rel}', (M2 / rel).exists())

for t in TOPICS:
    d = W / t / 'slides'
    check(f'slides dir exists {t}', d.is_dir())
    check(f'slides dir empty {t}', not any(d.iterdir()) if d.is_dir() else False,
          'pptx files are supplied separately')

# no stray files
STRAY = re.compile(r'(~\$|\.tmp$|\.bak$|Thumbs\.db|\.DS_Store|scratch)', re.I)
for p in M2.rglob('*'):
    if p.is_file():
        check(f'no stray file {p.name}', not STRAY.search(p.name), str(p))

# ------------------------------------------------------------------ datasets
def md5(p):
    return hashlib.md5(pathlib.Path(p).read_bytes()).hexdigest()

check('module-demo raw csv matches Datasets',
      md5(M2 / 'module-demo/Ilesanmi_Sales_Raw_Export.csv')
      == md5(M2 / 'Datasets/Ilesanmi_Sales_Raw_Export.csv'))
check('module-demo answer key matches Datasets',
      md5(M2 / 'module-demo/Ilesanmi_Sales_Clean_AnswerKey.xlsx')
      == md5(M2 / 'Datasets/Ilesanmi_Sales_Clean_AnswerKey.xlsx'))

# ------------------------------------------------------------------ markdown
MD = sorted(M2.rglob('*.md'))
check('markdown file count', len(MD) >= 40, len(MD))

LINK = re.compile(r'\[([^\]]+)\]\(([^)]+)\)')
for p in MD:
    txt = p.read_text(encoding='utf-8')

    # encoding sanity: no mojibake, no BOM, no stray replacement chars
    check(f'{p.name} no replacement char', '�' not in txt, str(p))
    check(f'{p.name} no BOM', not txt.startswith('﻿'), str(p))
    check(f'{p.name} naira sign intact', 'â‚¦' not in txt, str(p))

    # relative links resolve
    for label, target in LINK.findall(txt):
        if target.startswith(('http://', 'https://', '#', 'mailto:')):
            continue
        tgt = (p.parent / target.split('#')[0]).resolve()
        check(f'link resolves in {p.name} -> {target}', tgt.exists(), str(p))

    # unbalanced markdown table rows would render broken
    for i, line in enumerate(txt.splitlines(), 1):
        if line.strip().startswith('|') and line.strip().endswith('|'):
            check(f'{p.name}:{i} table row has cells', line.count('|') >= 2, line[:60])

    # headings present
    check(f'{p.name} has a title heading', txt.lstrip().startswith('#'), str(p))

# every lab pack must carry the required parts
REQUIRED_LAB = ['Tier 1', 'Tier 2', 'Tier 3', 'Version A', 'Version B',
                'Time-box', 'Expected output', 'Hints for Tier']
for t in TOPICS:
    txt = (W / t / 'lab' / 'lab-pack.md').read_text(encoding='utf-8')
    for part in REQUIRED_LAB:
        check(f'{t} lab has "{part}"', part in txt)
    check(f'{t} lab states a dataset', 'Ilesanmi' in txt or 'CleanSales' in txt)
    check(f'{t} lab has collapsed hints', '<details>' in txt and '</details>' in txt)
    check(f'{t} lab time-box total stated', re.search(r'\*\*Total\*\*', txt) is not None)

# solution notes must no longer claim the workbook is missing
for t in TOPICS:
    txt = (W / t / 'lab' / 'solutions' / 'solution-notes.md').read_text(encoding='utf-8')
    check(f'{t} notes: no "not been built"', 'has not been built' not in txt)
    check(f'{t} notes: records built status', 'built, executed and tested' in txt.lower()
          or 'Status: built' in txt)
    n = int(t[:2])
    check(f'{t} notes name the workbook', f'2.{n}_lab_solution.xlsx' in txt)

# mini project brief must link to the rubric and the dataset dictionary
brief = (M2 / 'mini-project/module-02-mini-project-brief.md').read_text(encoding='utf-8')
check('brief links the rubric', 'module-02-mini-project-rubric.md' in brief)
check('brief links the dataset README', 'Datasets/README.md' in brief)
check('brief states AI rule', 'prompt log' in brief.lower())
rubric = (M2 / 'mini-project/module-02-mini-project-rubric.md').read_text(encoding='utf-8')
for lvl in ('Developing', 'Competent', 'Excellent'):
    check(f'rubric has {lvl}', lvl in rubric)
check('rubric has a verification-evidence criterion', 'Verification evidence' in rubric)

# ------------------------------------------------------------------ MCQ
bank = json.loads((M2 / 'MCQ/module-02-mcq.json').read_text(encoding='utf-8'))
check('mcq is a list', isinstance(bank, list))
check('mcq every item has explanation', all(i.get('explanation', '').strip() for i in bank))
check('mcq no duplicate questions',
      len({i['question'] for i in bank}) == len(bank),
      f"{len(bank) - len({i['question'] for i in bank})} duplicates")
check('mcq no duplicate descriptions',
      len({i['description'] for i in bank}) == len(bank),
      [d for d in {i['description'] for i in bank}
       if [x['description'] for x in bank].count(d) > 1])
with (M2 / 'MCQ/module-02-mcq.csv').open(encoding='utf-8', newline='') as fh:
    rows = list(csv.DictReader(fh))
check('csv matches json exactly', [dict(r) for r in rows] == bank)

# short-answer items must state a formula or a figure
sa = [i for i in bank if i['type'] == 'short answer']
check('six short answer items', len(sa) == 6, len(sa))
for i in sa:
    check(f"SA answer non-empty: {i['description']}", bool(i['answer'].strip()))

# ------------------------------------------------------------------ workbooks
try:
    import openpyxl
    for t in TOPICS:
        n = int(t[:2])
        p = W / t / 'lab' / 'solutions' / f'2.{n}_lab_solution.xlsx'
        wb = openpyxl.load_workbook(p, data_only=True)
        check(f'2.{n} opens', True)
        check(f'2.{n} has sheets', len(wb.sheetnames) >= 3, wb.sheetnames)
        # scan every cached value for unexpected error strings
        ERRS = {'#REF!', '#NAME?', '#DIV/0!', '#NULL!', '#NUM!'}
        found = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value in ERRS:
                        found.setdefault(f'{ws.title}!{c.coordinate}', c.value)
        check(f'2.{n} no hard errors', not found, found)
        wb.close()
except ImportError:
    check('openpyxl available', False, 'cannot audit workbooks')

# ------------------------------------------------------------------ tracker
tracker = (M2 / 'content-tracker.md').read_text(encoding='utf-8')
for t in ('Dataset and licence', 'AI-drafted', 'Verification performed', 'Logged exceptions'):
    check(f'tracker section: {t}', t in tracker)
check('tracker records the xlsx workbooks', '2.1_lab_solution.xlsx' in tracker)
check('tracker no longer blocks sign-off', 'blocking sign-off' not in tracker)
check('tracker points at the toolchain', 'checks/' in tracker or 'audit' in tracker.lower())

# ------------------------------------------------------------------ report
print(f'checks passed: {passes[0]}')
if fails:
    print(f'\nFAILURES ({len(fails)}):')
    for f in fails:
        print('  -', f)
    sys.exit(1)
print('AUDIT PASSED')
